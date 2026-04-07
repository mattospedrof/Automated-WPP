import os, time, math, logging, threading, random, webbrowser
from datetime import datetime, timedelta
from tkinter import filedialog, messagebox

import customtkinter as ctk
import pandas as pd
import openpyxl

from src.bot import WhatsAppBot
from configs.conf_logs import conf_logging
from configs.config_colors import C, F
from views.sidebar   import sidebar
from views.main_view import main_view

conf_logging()

MESSAGE_TEMPLATE = ("")

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")

# @TAG: helper-first-name
def first_name(s):
    s = str(s).strip()
    return s.split()[0].capitalize() if s else "Cliente"

# @TAG: helper-build-message
def build_message(template, full_name):
    return template.replace("{primeiro_nome}", first_name(full_name))

# @TAG: helper-is-error
def is_erro(p):
    return str(p).strip().lower() in ("erro", "", "nan", "none")

# @TAG: helper-format-time
def fmt_time(s):
    if s < 0 or math.isinf(s) or math.isnan(s):
        return "--"
    td = timedelta(seconds=int(s))
    h, r = divmod(td.seconds, 3600)
    m, s = divmod(r, 60)
    if h: return f"{h}h {m:02d}min"
    if m: return f"{m}min {s:02d}s"
    return f"{s}s"

# @TAG: helper-error-file-path
def get_error_file_path():
    today = datetime.now().strftime("%d_%m_%Y")
    os.makedirs("tmp", exist_ok=True)
    return os.path.join("tmp", f"erros_{today}.txt")

# @TAG: helper-register-error
def register_error(cid, phone):
    with open(get_error_file_path(), "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now():%H:%M:%S}] ID: {cid} | Telefone: {phone}\n")


# @TAG: helper-phone-column
def get_phone_column(df) -> str:
    if "Celular" in df.columns:
        return "Celular"
    if "Telefone" in df.columns:
        return "Telefone"
    return None


# @TAG: sheet-updater-init
class SheetUpdater:
    def __init__(self, path):
        self.path = path
        self._lock = threading.Lock()
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        h = {c.value: c.column for c in ws[1]}
        self.col_status   = h.get("Status")
        self.col_mensagem = h.get("Mensagem")

        if not self.col_status:
            last_col = ws.max_column + 1
            ws.cell(1, last_col).value = "Status"
            self.col_status = last_col
        if not self.col_mensagem:
            last_col = ws.max_column + 1
            ws.cell(1, last_col).value = "Mensagem"
            self.col_mensagem = last_col

        wb.save(self.path)
        wb.close()

    # @TAG: sheet-updater-update
    def update_row(self, idx, status, mensagem=""):
        with self._lock:
            try:
                wb = openpyxl.load_workbook(self.path)
                ws = wb.active
                r = idx + 2
                if self.col_status:
                    ws.cell(r, self.col_status).value = status
                if self.col_mensagem and mensagem:
                    ws.cell(r, self.col_mensagem).value = mensagem
                wb.save(self.path)
            except Exception as e:
                logging.warning(f"Sheet update error: {e}")


# @TAG: app-init
class App(ctk.CTk):

    SB        = 280
    P         = 14
    H_FILE    = 120
    H_MID     = 470
    H_PROG    = 180
    H_LOG     = 300
    DELAY_MIN = 5.0
    DELAY_MAX = 10.0

    def __init__(self):
        super().__init__()
        self.title("Fkz Tech  •  WhatsApp Sender")
        self.configure(fg_color=C["bg"])
        self.after(50, lambda: self.wm_state("zoomed"))

        self.df         = None
        self.file_path  = None
        self.updater    = None
        self.bot        = None
        self.running    = False
        self._counters  = {"sent": 0, "not_found": 0, "skipped": 0, "error": 0}
        self._times     = []
        self._reset_job_id = None

        self._build_ui()

        self.protocol("WM_DELETE_WINDOW", self._on_close_app)

    # @TAG: app-build-ui
    def _build_ui(self):
        sidebar(self)
        main_view(self)

    # @TAG: app-close-handler
    def _on_close_app(self):
        if self.running:
            self.running = False
            self._log("⚠️ App fechando — interrompendo envio...")
            self.after(500, self._cleanup_and_exit)
        else:
            self._cleanup_and_exit()

    # @TAG: app-cleanup-exit
    def _cleanup_and_exit(self):
        if self.bot:
            try:
                self.bot.close()
            except Exception:
                pass
        self.destroy()

    # @TAG: app-load-file
    def _load_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("Todos", "*.*")])
        if not path:
            return
        try:
            self.df = (pd.read_csv(path) if path.endswith(".csv")
                       else pd.read_excel(path))
            self.file_path = path
            self.updater   = SheetUpdater(path) if not path.endswith(".csv") else None

            for col in ("Status", "Mensagem"):
                if col not in self.df.columns:
                    self.df[col] = None
                self.df[col] = self.df[col].astype(object)

            phone_col = get_phone_column(self.df)
            n      = len(self.df)
            n_erro = 0
            if phone_col:
                n_erro = (self.df[phone_col].astype(str)
                          .str.lower().isin(["erro", "", "nan", "none"])).sum()

            self.label_file.configure(
                text=f"✔   {os.path.basename(path)}",
                text_color=C["green"])
            self.label_preview_columns.configure(
                text=(f"{n} contatos  •  {n_erro} com 'Erro' (serão pulados)  •  "
                      f"Colunas: {', '.join(self.df.columns)}"))

            sugest_start = self._sugest_start_line()

            self.entry_range_start.delete(0, "end")
            self.entry_range_start.insert(0, str(sugest_start))
            self.entry_range_end.delete(0, "end")
            self.entry_range_end.insert(0, str(n))

            enviados = (self.df["Status"].astype(str).str.lower()
                        .isin(["enviado", "wpp n encontrado"])).sum()
            pendentes = n - enviados - n_erro
            self._log(f"📂 {os.path.basename(path)} — {n} linhas, {enviados} enviados, "
                      f"{pendentes} pendentes, {n_erro} com Erro")
            logging.info(f"Planilha carregada: {path} ({n} contatos, {enviados} ja enviados)")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível ler o arquivo:\n{e}")

    # @TAG: app-suggest-start
    def _sugest_start_line(self) -> int:
        if self.df is None or self.df.empty or "Status" not in self.df.columns:
            return 1
        status = self.df["Status"].astype(str).str.strip().str.lower()
        done_mask = status.isin(["enviado", "wpp n encontrado"])
        for idx, done in enumerate(done_mask):
            if not done:
                return idx + 1
        return 1

    # @TAG: app-help
    def _help(self):
        webbrowser.open_new_tab("https://seusite.com")

    # @TAG: app-connect
    def _connect(self):
        if self.bot:
            try:
                self.bot.close()
            except Exception:
                pass
            self.bot = None

        self.btn_connect.configure(state="disabled", text="Conectando...")
        self._log("🔄 Abrindo WhatsApp Web — escaneie o QR Code no navegador...")
        threading.Thread(target=self._do_connect, daemon=True).start()

    # @TAG: app-do-connect
    def _do_connect(self):
        try:
            has_session = WhatsAppBot.has_existing_session()
            if has_session:
                self.after(0, self._on_session_detected)

            self.bot = WhatsAppBot()
            self.bot.open_whatsapp()
            self.after(0, self._on_connected)
        except Exception as e:
            err = str(e).lower()
            if any(k in err for k in ("invalid session id", "session deleted",
                                      "disconnected", "not connected")):
                self.after(0, self._on_connect_cancelled)
            else:
                self.after(0, lambda: self._on_connect_error(str(e)))

    # @TAG: app-connected-callback
    def _on_connected(self):
        self.pill.configure(fg_color=C["green_bg"])
        self.lbl_status.configure(text="  ●  Conectado", text_color=C["green"])
        self.btn_connect.configure(state="disabled", text="✔  Conectado")
        self.btn_disconnect.configure(
            state="normal",
            text_color=C["danger"],
            fg_color=C["danger_bg"],
            border_color=C["danger"])
        self._log("✅ WhatsApp Web conectado!")
        logging.info("WhatsApp Web conectado.")

    # @TAG: app-session-detected
    def _on_session_detected(self):
        self._log("📂 Sessão existente detectada — pulando QR Code...")
        logging.info("Sessão existente detectada.")

    # @TAG: app-connect-cancelled
    def _on_connect_cancelled(self):
        if self.bot:
            try:
                self.bot.close()
            except Exception:
                pass
            self.bot = None
        self.btn_connect.configure(state="normal", text="Conectar WhatsApp")
        self._log("⚠️ Conexão cancelada — navegador fechado pelo usuário.")
        logging.info("Conexão cancelada pelo usuário.")

    # @TAG: app-connect-error
    def _on_connect_error(self, msg):
        if self.bot:
            try:
                self.bot.close()
            except Exception:
                pass
            self.bot = None
        self.btn_connect.configure(state="normal", text="Conectar WhatsApp")
        self._log(f"❌ Erro ao conectar.")
        logging.error(f"Erro de conexão: {msg}")
        messagebox.showerror("Erro de conexão",
            "Não foi possível conectar ao WhatsApp Web.\n\n"
            "Verifique sua conexão e tente novamente.")

    # @TAG: app-disconnect
    def _disconnect(self):
        if self.bot:
            self.bot.close()
            self.bot = None
        try:
            import shutil
            session_dir = "./wa_session"
            if os.path.isdir(session_dir):
                shutil.rmtree(session_dir, ignore_errors=True)
                self._log("🗑️ Sessão removida.")
        except Exception:
            pass
        self.pill.configure(fg_color=C["danger_bg"])
        self.lbl_status.configure(text="  ●  Desconectado", text_color=C["danger"])
        self.btn_connect.configure(state="normal", text="Conectar WhatsApp")
        self.btn_disconnect.configure(
            state="disabled",
            text_color=C["muted"],
            fg_color=C["card2"],
            border_color=C["border"])
        self._log("🔌 Desconectado.")

    # @TAG: app-start-send
    def _start(self):
        if not self.bot:
            messagebox.showwarning("Aviso", "Conecte ao WhatsApp primeiro!"); return
        if self.df is None or self.df.empty:
            messagebox.showwarning("Aviso", "Carregue uma planilha de contatos!"); return

        phone_col = get_phone_column(self.df)
        if not phone_col:
            messagebox.showerror("Erro", "Coluna 'Celular' ou 'Telefone' não encontrada."); return
        if "Cliente" not in self.df.columns:
            messagebox.showerror("Erro", f"Coluna 'Cliente' não encontrada."); return

        template = self.txt_msg.get("0.0", "end").strip()
        if not template:
            messagebox.showwarning("Aviso",
                "O campo de mensagem está vazio!\n"
                "Cole ou escreva o texto antes de iniciar."); return
        try:
            delay_min   = App.DELAY_MIN
            delay_max   = App.DELAY_MAX
            range_start = int(self.entry_range_start.get()) - 1
            range_end   = int(self.entry_range_end.get())
        except ValueError:
            messagebox.showerror("Erro", "Verifique os valores de configuração."); return
        if range_start < 0 or range_end > len(self.df) or range_start >= range_end:
            messagebox.showerror(
                "Erro",
                f"Intervalo inválido. A planilha tem {len(self.df)} linhas."); return

        self.df["_phone_normalized"] = self.df[phone_col].apply(
            lambda x: WhatsAppBot.normalize_phone(str(x)))
        n_invalid = self.df["_phone_normalized"].isna().sum()
        if n_invalid > 0:
            self._log(f"⚠️ {n_invalid} telefone(s) invalido(s) — serao pulados")

        if self._reset_job_id:
            self.after_cancel(self._reset_job_id)
            self._reset_job_id = None

        self.running   = True
        self._counters = {"sent": 0, "not_found": 0, "skipped": 0, "error": 0}
        self._times    = []
        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.progress_bar.set(0)
        self.lbl_sent_total.configure(text="0 / ?")
        self.lbl_eta.configure(text="Tempo restante:  calculando...")
        threading.Thread(
            target=self._send_loop,
            args=(template, delay_min, delay_max, range_start, range_end),
            daemon=True).start()

    # @TAG: app-send-loop-wrapper
    def _send_loop(self, template, dmin, dmax, rs, re):
        try:
            self._send_loop_inner(template, dmin, dmax, rs, re)
        except Exception as exc:
            import traceback
            tb = traceback.format_exc()
            logging.error(f"Erro fatal:\n{tb}")
            self._log(f"\n💥 ERRO FATAL:\n{tb}")
            self.after(0, lambda: messagebox.showerror(
                "Erro no envio", f"Envio interrompido:\n\n{exc}"))
            self.after(0, self._reset_buttons)

    # @TAG: app-reset-buttons
    def _reset_buttons(self):
        self.running = False
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        if self._reset_job_id:
            self.after_cancel(self._reset_job_id)
        self._reset_job_id = self.after(10000, self._reset_after_completion)

    # @TAG: app-send-loop-core
    def _send_loop_inner(self, template, dmin, dmax, rs, re):
        slice_df    = self.df.iloc[rs:re]
        already_done = slice_df["Status"].astype(str).str.strip().str.lower().isin(
                           ["enviado", "wpp n encontrado"])
        phone_valid  = ~slice_df["_phone_normalized"].isna()
        valid_mask   = phone_valid & ~already_done
        total_valid  = int(valid_mask.sum())
        processed    = 0

        n_skipped_done = int(already_done.sum())

        self._log(f"🚀 Linhas {rs+1}–{re}  ({len(slice_df)} intervalo, {total_valid} pendentes, "
                  f"{n_skipped_done} ja enviados)")
        logging.info(f"Envio iniciado [{rs+1},{re}] — {total_valid} pendentes, {n_skipped_done} ja enviados")

        for pidx, row in slice_df.iterrows():
            if not self.running:
                self._log("⏹ Interrompido."); break

            if self.bot and not self.bot.is_connected():
                self._log("⚠️ Bot desconectado durante o envio!")
                logging.warning("Bot desconectado durante o envio.")
                self.after(0, lambda: messagebox.showwarning(
                    "Desconectado",
                    "O WhatsApp Web foi fechado ou desconectou.\n\n"
                    "O envio foi interrompido. Reconecte e tente novamente."))
                self.after(0, self._reset_buttons)
                return

            phone = row["_phone_normalized"]
            status_atual = str(row.get("Status", "")).strip().lower()

            cid   = row["Id"]            if "Id"       in row.index else "?"
            fname = str(row["Cliente"])  if "Cliente"  in row.index else "Cliente"

            if status_atual in ("enviado", "wpp n encontrado"):
                self._log(f"⏭ ID:{cid}  Pulado (ja enviado: {row['Status']})")
                logging.info(f"Pulado | ID:{cid} | ja enviado: {row['Status']}")
                self.after(0, self._update_counters)
                continue

            if phone is None:
                self._counters["skipped"] += 1
                self._log(f"⏭ ID:{cid}  Pulado (Telefone invalido)")
                logging.info(f"Pulado | ID:{cid} | telefone invalido")
                self.after(0, self._update_counters)
                continue

            t0  = time.time()
            msg = build_message(template, fname)
            self._log(f"\n[{processed+1}/{total_valid}]  ID:{cid}  {first_name(fname)}  {phone}")
            logging.info(f"Processando ID:{cid} | {phone}")

            result = self.bot.send_message(phone, msg)

            if result == "sent":
                status = "Enviado"
                self._log("✅ Enviado")
                self._counters["sent"] += 1
            elif result == "wpp_not_found":
                status = "wpp n encontrado"
                self._log("🟠 Não encontrado")
                self._counters["not_found"] += 1
            else:
                status = "Erro"
                self._log("❌ Erro")
                self._counters["error"] += 1
                register_error(cid, phone)

            logging.info(f"{status} | ID:{cid} | {phone}")

            if self.updater:
                self.updater.update_row(pidx, status, msg if result == "sent" else "")
            if "Status" in self.df.columns:
                self.df.at[pidx, "Status"] = status
            if "Mensagem" in self.df.columns and result == "sent":
                self.df.at[pidx, "Mensagem"] = msg

            delay = random.uniform(dmin, dmax)
            self._log(f"   ⏱ {delay:.1f}s...")
            time.sleep(delay)

            elapsed = time.time() - t0
            self._times.append(elapsed)
            processed += 1
            remaining = total_valid - processed
            avg       = sum(self._times) / len(self._times)
            eta       = avg * remaining
            prog      = processed / total_valid if total_valid else 0

            self.after(0, self.progress_bar.set, prog)
            self.after(0, self._update_sent_total, processed, total_valid)
            self.after(0, self.lbl_eta.configure, {
                "text": f"Tempo restante:  {fmt_time(eta)}   •   média/msg: {fmt_time(avg)}"})
            self.after(0, self._update_counters)

        self.after(0, self._on_done, processed, total_valid)

    # @TAG: app-send-complete
    def _on_done(self, processed, total_valid):
        self.running = False
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        self.lbl_eta.configure(text="Tempo restante:  --   •   ✅ Concluído")
        c = self._counters
        summary = (f"Envio finalizado!\n\n"
                   f"✅ Enviados:         {c['sent']}\n"
                   f"🟠 Não encontrados:  {c['not_found']}\n"
                   f"⏭ Pulados (Erro):   {c['skipped']}\n"
                   f"❌ Erros:           {c['error']}\n\n"
                   f"Verifique sua planilha!")
        self._log(f"\n🏁 Concluído!  ✅{c['sent']}  🟠{c['not_found']}  "
                  f"⏭{c['skipped']}  ❌{c['error']}")
        logging.info(f"Concluído. {c}")
        messagebox.showinfo("Concluído", summary)

        self._reset_job_id = self.after(10000, self._reset_after_completion)

    # @TAG: app-auto-reset
    def _reset_after_completion(self):
        if self.running:
            return
        self._reset_job_id = None

        self.df = None
        self.file_path = None
        self.updater = None

        self.label_file.configure(
            text="Nenhum arquivo selecionado - Consulte no botão de ajuda como o arquivo deve estar formatado",
            text_color=C["muted"])
        self.label_preview_columns.configure(text="")

        self.entry_range_start.delete(0, "end")
        self.entry_range_start.insert(0, "1")
        self.entry_range_end.delete(0, "end")
        self.entry_range_end.insert(0, "0")

        self.progress_bar.set(0)
        self.lbl_sent_total.configure(text="0 / ?")
        self.lbl_eta.configure(text="Tempo restante:  --")

        self._counters = {"sent": 0, "not_found": 0, "skipped": 0, "error": 0}
        self.after(0, self._update_counters)

        self._log("🔄 Campos resetados — pronto para novo envio.")

    # @TAG: app-stop
    def _stop(self):
        self.running = False
        self.btn_stop.configure(state="disabled")
        self._log("⏹ Envio parado pelo usuário.")
        if self._reset_job_id:
            self.after_cancel(self._reset_job_id)
        self._reset_job_id = self.after(10000, self._reset_after_completion)

    # @TAG: app-log
    def _log(self, text: str):
        def _do():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", text + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _do)

    # @TAG: app-clear-log
    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("0.0", "end")
        self.log_box.configure(state="disabled")

    # @TAG: app-ui-sent-total
    def _update_sent_total(self, sent, total):
        self.lbl_sent_total.configure(text=f"{sent} / {total}")

    # @TAG: app-ui-counters
    def _update_counters(self):
        c = self._counters
        self.cnt_sent.configure(text=str(c["sent"]))
        self.cnt_nfound.configure(text=str(c["not_found"]))
        self.cnt_skipped.configure(text=str(c["skipped"]))
        self.cnt_error.configure(text=str(c["error"]))


if __name__ == "__main__":
    app = App()
    app.mainloop()
